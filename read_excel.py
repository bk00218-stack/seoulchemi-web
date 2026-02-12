import openpyxl
import sys

wb = openpyxl.load_workbook(sys.argv[1])
for sheet_name in wb.sheetnames:
    print(f"=== Sheet: {sheet_name} ===")
    ws = wb[sheet_name]
    for i, row in enumerate(ws.iter_rows(max_row=30, values_only=True)):
        if any(cell is not None for cell in row):
            print(list(row))
