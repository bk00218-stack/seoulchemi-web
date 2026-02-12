# -*- coding: utf-8 -*-
import openpyxl
import json
import sys

# Load the Excel file
wb = openpyxl.load_workbook(sys.argv[1])
ws = wb['거래처List']

# Get headers from first row
headers = [cell.value for cell in ws[1]]
print("Headers:", headers)

# Column mapping (0-indexed)
# Based on the data: No, 상호, 대표자, 사업자번호, 전화번호, HP, (주소1), (주소2), 업태, 업종, 청구일, 마이선스, 초기잔액, 가격정책, 배송담당, E-Mail, 등록일, 거래처유형
COL_NO = 0
COL_NAME = 1
COL_OWNER = 2
COL_BIZ_NO = 3
COL_PHONE = 4
COL_HP = 5
COL_ADDR1 = 6
COL_ADDR2 = 7
COL_BIZ_TYPE = 8
COL_BIZ_CAT = 9
COL_BILLING_DAY = 10
COL_LICENSE = 11
COL_INIT_BALANCE = 12
COL_PRICE_POLICY = 13
COL_DELIVERY = 14
COL_EMAIL = 15
COL_REG_DATE = 16
COL_STORE_TYPE = 17

stores = []

for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    # Skip empty rows
    if not row[COL_NAME]:
        continue
    
    name = str(row[COL_NAME]).strip() if row[COL_NAME] else None
    if not name:
        continue
    
    # Parse phone - clean up
    phone = str(row[COL_PHONE]).strip() if row[COL_PHONE] else None
    if phone and phone != '-':
        phone = phone.replace(' ', '-')
    else:
        phone = None
    
    # Parse address
    addr1 = str(row[COL_ADDR1]).strip() if row[COL_ADDR1] else ''
    addr2 = str(row[COL_ADDR2]).strip() if row[COL_ADDR2] else ''
    address = f"{addr1} {addr2}".strip() if addr1 or addr2 else None
    
    # Parse billing day
    billing_day = None
    if row[COL_BILLING_DAY]:
        try:
            billing_day = int(row[COL_BILLING_DAY])
        except:
            pass
    
    # Parse initial balance (미수금)
    outstanding = 0
    if row[COL_INIT_BALANCE]:
        try:
            outstanding = int(float(row[COL_INIT_BALANCE]))
        except:
            pass
    
    store = {
        'code': str(row[COL_NO]) if row[COL_NO] else None,
        'name': name,
        'ownerName': str(row[COL_OWNER]).strip() if row[COL_OWNER] else None,
        'businessRegNo': str(row[COL_BIZ_NO]).strip() if row[COL_BIZ_NO] else None,
        'phone': phone,
        'address': address,
        'businessType': str(row[COL_BIZ_TYPE]).strip() if row[COL_BIZ_TYPE] else None,
        'businessCategory': str(row[COL_BIZ_CAT]).strip() if row[COL_BIZ_CAT] else None,
        'billingDay': billing_day,
        'outstandingAmount': outstanding,
        'storeType': str(row[COL_STORE_TYPE]).strip() if row[COL_STORE_TYPE] else None,
        'email': str(row[COL_EMAIL]).strip() if row[COL_EMAIL] else None,
        'areaCode': str(row[COL_DELIVERY]).strip() if row[COL_DELIVERY] else None,
        'status': 'active',
        'isActive': True,
    }
    
    # Clean up None strings
    for key in store:
        if store[key] == 'None' or store[key] == '-':
            store[key] = None
    
    stores.append(store)

print(f"\nTotal stores parsed: {len(stores)}")
print(f"\nFirst 3 stores:")
for s in stores[:3]:
    print(json.dumps(s, ensure_ascii=False, indent=2))

# Output to JSON file
with open('stores_import.json', 'w', encoding='utf-8') as f:
    json.dump(stores, f, ensure_ascii=False, indent=2, default=str)

print(f"\nData exported to stores_import.json")
